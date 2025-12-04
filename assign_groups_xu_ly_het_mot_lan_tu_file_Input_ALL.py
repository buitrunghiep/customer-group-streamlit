#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
assign_groups_to_excel.py
-------------------------
Äá»c file Excel "INPUT ALL.xlsx" (hoáº·c file chá»‰ Ä‘á»‹nh), cÃ³ 3 sheet:
 - Customers (cá»™t: CustomerID, TypeOfCustomer)
 - GroupName (cá»™t: GroupName) - danh sÃ¡ch cÃ¡c group há»£p lá»‡
 - GroupSize (cá»™t: GroupName, TypeOfCustomer, Size) - quota theo loáº¡i KH cho tá»«ng group
Sau Ä‘Ã³ xáº¿p khÃ¡ch hÃ ng vÃ o group sao cho Ä‘Ã¡p á»©ng Ä‘Ãºng quota theo loáº¡i khÃ¡ch hÃ ng,
vÃ  ghi thÃªm 1 sheet má»›i "Assigned" vÃ o CHÃNH file Excel nguá»“n.

CÃ¡ch dÃ¹ng:
    python assign_groups - LÃ m háº¿t má»™t láº§n tá»« file Input ALL.py --input "20251126 - INPUT ALL.xlsx" [--seed 42]

YÃªu cáº§u:
    pip install pandas openpyxl numpy
"""
import argparse
import sys
import os
from typing import List
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def read_input(xlsx_path: str):
    try:
        customers = pd.read_excel(xlsx_path, sheet_name="Customers")
    except Exception as e:
        raise RuntimeError(f"KhÃ´ng Ä‘á»c Ä‘Æ°á»£c sheet 'Customers'. Chi tiáº¿t: {e}")
    try:
        groups = pd.read_excel(xlsx_path, sheet_name="GroupName")
    except Exception as e:
        raise RuntimeError(f"KhÃ´ng Ä‘á»c Ä‘Æ°á»£c sheet 'GroupName'. Chi tiáº¿t: {e}")
    try:
        groupsize = pd.read_excel(xlsx_path, sheet_name="GroupSize")
    except Exception as e:
        raise RuntimeError(f"KhÃ´ng Ä‘á»c Ä‘Æ°á»£c sheet 'GroupSize'. Chi tiáº¿t: {e}")
    return customers, groups, groupsize

def validate(customers: pd.DataFrame, groups: pd.DataFrame, groupsize: pd.DataFrame):
    # Basic columns
    for col in ["CustomerID", "TypeOfCustomer"]:
        if col not in customers.columns:
            raise ValueError(f"Sheet Customers thiáº¿u cá»™t báº¯t buá»™c: '{col}'.")

    if "GroupName" not in groups.columns:
        raise ValueError("Sheet GroupName pháº£i cÃ³ cá»™t 'GroupName'.")

    for col in ["GroupName", "TypeOfCustomer", "Size"]:
        if col not in groupsize.columns:
            raise ValueError("Sheet GroupSize pháº£i cÃ³ Ä‘á»§ cá»™t: 'GroupName', 'TypeOfCustomer', 'Size'.")

    # Clean
    customers = customers.dropna(how="all").copy()
    groups = groups.dropna(how="all").copy()
    groupsize = groupsize.dropna(how="all").copy()

    if customers.empty:
        raise ValueError("Sheet Customers khÃ´ng cÃ³ dá»¯ liá»‡u.")
    if groups.empty:
        raise ValueError("Sheet GroupName khÃ´ng cÃ³ dá»¯ liá»‡u.")
    if groupsize.empty:
        raise ValueError("Sheet GroupSize khÃ´ng cÃ³ dá»¯ liá»‡u.")

    # Normalize types
    customers["CustomerID"] = customers["CustomerID"].astype(str).str.strip()
    customers["TypeOfCustomer"] = customers["TypeOfCustomer"].astype(str).str.strip()
    groups["GroupName"] = groups["GroupName"].astype(str).str.strip()
    groupsize["GroupName"] = groupsize["GroupName"].astype(str).str.strip()
    groupsize["TypeOfCustomer"] = groupsize["TypeOfCustomer"].astype(str).str.strip()
    groupsize["Size"] = pd.to_numeric(groupsize["Size"], errors="coerce").fillna(0).astype(int)

    if (groups["GroupName"] == "").any():
        raise ValueError("Sheet GroupName cÃ³ GroupName rá»—ng.")
    if (groupsize["GroupName"] == "").any() or (groupsize["TypeOfCustomer"] == "").any():
        raise ValueError("Sheet GroupSize cÃ³ GroupName/TypeOfCustomer rá»—ng.")

    # Ensure all group names used in GroupSize are valid in GroupName sheet
    valid_groups = set(groups["GroupName"].unique().tolist())
    used_groups = set(groupsize["GroupName"].unique().tolist())
    invalid = used_groups - valid_groups
    if invalid:
        raise ValueError(f"CÃ³ GroupName trong GroupSize khÃ´ng náº±m trong sheet GroupName: {sorted(invalid)}")

    # Check per-type totals equal availability
    need_per_type = groupsize.groupby("TypeOfCustomer")["Size"].sum()
    have_per_type = customers["TypeOfCustomer"].value_counts()

    errs = []
    for t, need in need_per_type.items():
        have = int(have_per_type.get(t, 0))
        if have != int(need):
            errs.append(f"- Loáº¡i '{t}': cáº§n {int(need)}, cÃ³ {have}")
    # Also ensure no extra types in Customers missing from GroupSize
    for t, have in have_per_type.items():
        need = int(need_per_type.get(t, 0))
        if have != need:
            if f"- Loáº¡i '{t}': cáº§n {need}, cÃ³ {int(have)}" not in errs:
                errs.append(f"- Loáº¡i '{t}': cáº§n {need}, cÃ³ {int(have)}")
    if errs:
        raise ValueError("Tá»•ng theo Loáº¡i khÃ´ng khá»›p giá»¯a Customers vÃ  GroupSize:\n" + "\n".join(errs))

    return customers, groups, groupsize

def assign(customers: pd.DataFrame, groups: pd.DataFrame, groupsize: pd.DataFrame, seed: int = None) -> pd.DataFrame:
    rng = np.random.default_rng(seed)
    group_order = groups["GroupName"].tolist()

    # Build for each type: list of group labels respecting quotas
    labels_by_type = {}
    for t, subq in groupsize.groupby("TypeOfCustomer"):
        labels = []
        # respect group order from GroupName sheet
        subq_by_order = {g:0 for g in group_order}
        for _, r in subq.iterrows():
            subq_by_order[r["GroupName"]] = subq_by_order.get(r["GroupName"], 0) + int(r["Size"])
        for g in group_order:
            q = subq_by_order.get(g, 0)
            if q > 0:
                labels.extend([g]*q)
        labels_by_type[t] = labels

    # For each type: shuffle customers of that type and assign labels
    assigned_parts = []
    for t, labels in labels_by_type.items():
        sub = customers[customers["TypeOfCustomer"] == t].copy().reset_index(drop=True)
        n = len(sub)
        if len(labels) != n:
            raise RuntimeError(f"Quota cá»§a loáº¡i '{t}' khÃ´ng khá»›p sá»‘ KH: labels={len(labels)} n={n}")
        perm = rng.permutation(n)
        sub = sub.iloc[perm].reset_index(drop=True)
        sub["Group"] = labels
        assigned_parts.append(sub)

    assigned = pd.concat(assigned_parts, ignore_index=True)
    # Optional: shuffle overall order
    assigned = assigned.sample(frac=1.0, random_state=rng.integers(0, 2**32-1)).reset_index(drop=True)
    return assigned

def write_assigned_to_same_file(xlsx_path: str, assigned: pd.DataFrame, sheet_name: str = "Assigned"):
    wb = load_workbook(xlsx_path)
    if sheet_name in wb.sheetnames:
        # remove old sheet to replace
        ws_old = wb[sheet_name]
        wb.remove(ws_old)
    ws = wb.create_sheet(title=sheet_name)
    # write DataFrame
    for r in dataframe_to_rows(assigned, index=False, header=True):
        ws.append(r)
    wb.save(xlsx_path)

def main():
    parser = argparse.ArgumentParser(description="Xáº¿p khÃ¡ch hÃ ng vÃ o Group theo quota tá»«ng loáº¡i (ghi sheet má»›i vÃ o chÃ­nh file).")
    parser.add_argument("--input", "-i", required=True, help="ÄÆ°á»ng dáº«n file Excel nguá»“n (20251126 - INPUT ALL.xlsx).")
    parser.add_argument("--seed", type=int, default=None, help="Seed ngáº«u nhiÃªn (tuá»³ chá»n).")
    parser.add_argument("--sheet", default="Assigned", help="TÃªn sheet output (máº·c Ä‘á»‹nh: Assigned).")
    args = parser.parse_args()

    customers, groups, groupsize = read_input(args.input)
    customers, groups, groupsize = validate(customers, groups, groupsize)
    assigned = assign(customers, groups, groupsize, seed=args.seed)
    write_assigned_to_same_file(args.input, assigned, sheet_name=args.sheet)
    print(f"OK: ÄÃ£ ghi sheet '{args.sheet}' vÃ o file: {args.input} (tá»•ng {len(assigned)} khÃ¡ch hÃ ng).")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"Lá»–I: {e}", file=sys.stderr)
        sys.exit(1)